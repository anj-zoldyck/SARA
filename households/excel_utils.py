"""
Excel utilities for RBI Form A (Revised 2024) import/export.
Handles parsing and generation of Excel files matching the barangay spreadsheet layout.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime, date
from households.models import RELATIONSHIP_CHOICES, CIVIL_STATUS_CHOICES, SEX_CHOICES


class RBIFormAImporter:
    """
    Parses RBI Form A Excel spreadsheets for bulk household/family member import.
    
    Expected layout:
    Header block: Region, Province, City/Municipality, Barangay, Household Address, No. of Household Members
    Table columns: LAST NAME, FIRST NAME, MIDDLE NAME, EXT, PLACE OF BIRTH, DATE OF BIRTH, AGE, SEX,
                   CIVIL STATUS, CITIZENSHIP, OCCUPATION, Indicate if Labor/employed, Unemployed, PWD, OFW,
                   Solo Parent (OSY), Out of School Children (OSC) and/or IP
    """
    
    COLUMN_MAPPING = {
        'LAST NAME': 'last_name',
        'FIRST NAME': 'first_name',
        'MIDDLE NAME': 'middle_name',
        'EXT': 'suffix',
        'PLACE OF BIRTH': 'birthplace',
        'DATE OF BIRTH': 'birthdate',
        'AGE': 'age',
        'SEX': 'sex',
        'CIVIL STATUS': 'civil_status',
        'CITIZENSHIP': 'citizenship',
        'OCCUPATION': 'occupation',
        'INDICATE IF': 'indicate_if',
    }
    
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None
        self.header_data = {}
        self.members_data = []
        self.warnings = []
        self.errors = []
    
    def parse(self):
        """Parse the Excel file and extract header data and member rows."""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.worksheet = self.workbook.active
            
            self._parse_header_block()
            self._parse_member_table()
            
            return True, self.warnings
        except Exception as e:
            self.errors.append(f"Failed to parse Excel file: {str(e)}")
            return False, self.errors
        finally:
            if self.workbook is not None:
                self.workbook.close()
    
    def _parse_header_block(self):
        """Extract header information from the spreadsheet."""
        # Header block typically in rows 1-6
        # Look for key header fields
        header_keywords = ['Region', 'Province', 'City/Municipality', 'Barangay', 
                          'Household Address', 'No. of Household Members']
        
        for row_idx in range(1, 10):  # Scan wider range to capture all header fields
            for col_idx in range(1, 10):
                cell_value = self.worksheet.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    cell_str = str(cell_value).strip()
                    for keyword in header_keywords:
                        if keyword.lower() in cell_str.lower():
                            # Try to get the value from the next cell
                            next_cell = self.worksheet.cell(row=row_idx, column=col_idx + 1).value
                            if next_cell:
                                self.header_data[keyword] = str(next_cell).strip()
                            else:
                                # Value might be in the same cell after a colon
                                if ':' in cell_str:
                                    parts = cell_str.split(':', 1)
                                    if len(parts) > 1:
                                        self.header_data[keyword] = parts[1].strip()
    
    def _parse_member_table(self):
        """Parse the member table rows."""
        # Find the header row (where column names are)
        header_row = self._find_table_header_row()
        if not header_row:
            self.errors.append("Could not find member table header row")
            return
        
        # Map column indices to field names
        col_mapping = {}
        for col_idx in range(1, 20):
            cell_value = self.worksheet.cell(row=header_row, column=col_idx).value
            if cell_value:
                cell_str = str(cell_value).strip().upper()
                for col_name, field_name in self.COLUMN_MAPPING.items():
                    if col_name in cell_str:
                        col_mapping[col_idx] = field_name
                        break
        
        # Parse member rows starting from the next row
        row_idx = header_row + 1
        while True:
            row_data = self._parse_member_row(row_idx, col_mapping)
            if row_data is None:  # Empty row or end of data
                break
            if row_data:  # Valid member data
                self.members_data.append(row_data)
            row_idx += 1
        
        # Assign default relationships (RBI Form A doesn't collect relationship data)
        for idx, member in enumerate(self.members_data):
            if idx == 0:
                member['relationship'] = 'HEAD'  # First member is head of family
            else:
                member['relationship'] = 'OTHER_RELATIVE'  # Safe generic default
    
    def _find_table_header_row(self):
        """Find the row containing the table column headers."""
        for row_idx in range(1, 20):
            first_cell = str(self.worksheet.cell(row=row_idx, column=1).value or '').strip().upper()
            if 'LAST NAME' in first_cell:
                return row_idx
        return None
    
    def _parse_member_row(self, row_idx, col_mapping):
        """Parse a single member row."""
        row_values = {}
        has_data = False

        for col_idx, field_name in col_mapping.items():
            cell_value = self.worksheet.cell(row=row_idx, column=col_idx).value
            if cell_value is not None and cell_value != '':
                has_data = True
                # Preserve native types for fields that need them (dates, numbers).
                # Only stringify+strip for genuinely text fields.
                if field_name in ('birthdate', 'age'):
                    row_values[field_name] = cell_value  # keep as datetime/date/int/float
                else:
                    cell_str = str(cell_value).strip()
                    if cell_str:
                        row_values[field_name] = cell_str

        if not has_data:
            return None

        # Process and validate the row
        return self._process_member_row(row_values, row_idx)
    
    def _process_member_row(self, row_values, row_idx):
        """Process and validate a member row."""
        processed = {
            'row_number': row_idx,
            'last_name': row_values.get('last_name', ''),
            'first_name': row_values.get('first_name', ''),
            'middle_name': row_values.get('middle_name', ''),
            'suffix': row_values.get('suffix', ''),
            'birthplace': row_values.get('birthplace', ''),
            'birthdate': self._parse_date(row_values.get('birthdate')),
            'age': self._parse_age(row_values.get('age')),
            'sex': self._parse_sex(row_values.get('sex')),
            'civil_status': self._parse_civil_status(row_values.get('civil_status')),
            'citizenship': row_values.get('citizenship', 'Filipino'),
            'occupation': row_values.get('occupation', ''),
            'indicate_if': row_values.get('indicate_if', ''),
        }
        
        # Parse the "Indicate if" column into boolean flags
        self._parse_indicate_if(processed)
        
        # Validate age vs birthdate
        self._validate_age_birthdate(processed, row_idx)
        
        # Check required fields
        self._validate_required_fields(processed, row_idx)
        
        return processed
    
    def _parse_date(self, date_str):
        """Parse a date string into a date object."""
        if not date_str:
            return None

        # Check datetime FIRST since datetime is a subclass of date —
        # checking `date` first would incorrectly catch datetime objects too.
        if isinstance(date_str, datetime):
            return date_str.date()

        if isinstance(date_str, date):
            return date_str

        # Try various date formats
        date_formats = ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d']
        for fmt in date_formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except (ValueError, TypeError):
                continue

        # Try Excel serial date
        try:
            if isinstance(date_str, (int, float)):
                return datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(date_str) - 2).date()
        except:
            pass

        return None
    
    def _parse_age(self, age_str):
        """Parse age string to integer."""
        if not age_str:
            return None
        try:
            # Handle both string and numeric (int/float) inputs
            # Excel may store age as float (e.g., 25.0)
            if isinstance(age_str, (int, float)):
                return int(age_str)
            return int(str(age_str).strip())
        except (ValueError, TypeError):
            return None
    
    def _parse_sex(self, sex_str):
        """Parse sex value to match model choices."""
        if not sex_str:
            return None
        
        sex_str = str(sex_str).strip().upper()
        if sex_str in ['M', 'MALE', 'M']:
            return 'M'
        elif sex_str in ['F', 'FEMALE', 'F']:
            return 'F'
        return None
    
    def _parse_civil_status(self, civil_status_str):
        """Parse civil status to match model choices."""
        if not civil_status_str:
            return None
        
        civil_status_str = str(civil_status_str).strip().title()
        for choice_value, choice_label in CIVIL_STATUS_CHOICES:
            if choice_label == civil_status_str or choice_value == civil_status_str.upper():
                return choice_value
        return None
    
    def _parse_indicate_if(self, processed):
        """Parse the 'Indicate if' column into boolean flags."""
        indicate_if = str(processed.get('indicate_if', '')).lower()
        
        processed['is_pwd'] = 'pwd' in indicate_if
        processed['is_solo_parent'] = 'solo parent' in indicate_if
        processed['is_out_of_school_youth'] = 'osy' in indicate_if or 'out of school youth' in indicate_if
        processed['is_out_of_school_children'] = 'osc' in indicate_if or 'out of school children' in indicate_if
        processed['is_indigenous'] = 'ip' in indicate_if or 'indigenous' in indicate_if
        processed['is_senior_citizen'] = 'senior' in indicate_if or 'senior citizen' in indicate_if
        
        # Check for unparsed text
        parsed_keywords = ['pwd', 'solo parent', 'osy', 'out of school youth', 'osc', 'out of school children', 'ip', 'indigenous', 'senior', 'senior citizen']
        remaining_text = indicate_if
        for keyword in parsed_keywords:
            remaining_text = remaining_text.replace(keyword, '').strip()
        
        if remaining_text and remaining_text not in ['', ',', '/', '-', 'and', 'or']:
            processed['unparsed_indicate_if'] = remaining_text
    
    def _validate_age_birthdate(self, processed, row_idx):
        """Validate that age matches birthdate."""
        birthdate = processed.get('birthdate')
        age = processed.get('age')
        
        if birthdate and age:
            calculated_age = date.today().year - birthdate.year - ((date.today().month, date.today().day) < (birthdate.month, birthdate.day))
            if abs(calculated_age - age) > 1:  # Allow 1 year tolerance
                self.warnings.append({
                    'row': row_idx,
                    'field': 'age/birthdate',
                    'message': f"Age ({age}) doesn't match birthdate ({birthdate}). Calculated age: {calculated_age}"
                })
    
    def _validate_required_fields(self, processed, row_idx):
        """Validate required fields for FamilyMember."""
        required_fields = ['last_name', 'first_name', 'sex']
        missing_fields = []
        
        for field in required_fields:
            if not processed.get(field):
                missing_fields.append(field)
        
        if missing_fields:
            self.errors.append({
                'row': row_idx,
                'field': ', '.join(missing_fields),
                'message': f"Missing required fields: {', '.join(missing_fields)}"
            })
            processed['has_errors'] = True
    
    def get_header_data(self):
        """Return parsed header data."""
        return self.header_data
    
    def get_members_data(self):
        """Return parsed member data with dates converted to ISO strings for JSON serialization."""
        # Convert date objects to ISO strings for session storage
        members_copy = []
        for member in self.members_data:
            member_copy = member.copy()
            if member_copy.get('birthdate') and isinstance(member_copy['birthdate'], date):
                member_copy['birthdate'] = member_copy['birthdate'].isoformat()
            members_copy.append(member_copy)
        return members_copy
    
    def get_warnings(self):
        """Return parsing warnings."""
        return self.warnings
    
    def get_errors(self):
        """Return parsing errors."""
        return self.errors


class RBIFormAExporter:
    """
    Generates RBI Form A Excel spreadsheets for export.
    """
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
    
    def generate_single_household(self, household, family):
        """Generate an Excel file for a single household/family."""
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "RBI Form A"
        
        # Write header block
        self._write_header_block(household, family)
        
        # Write table headers
        self._write_table_headers()
        
        # Write member rows
        self._write_member_rows(family)
        
        # Apply styling
        self._apply_styling()
        
        return self.workbook
    
    def generate_multiple_households(self, households):
        """Generate an Excel file with multiple households (one per sheet)."""
        self.workbook = openpyxl.Workbook()
        # Remove default sheet
        self.workbook.remove(self.workbook.active)
        
        for idx, household in enumerate(households):
            # Get the first/only family for this household
            family = household.families.filter(is_active=True).first()
            if not family:
                continue
            
            # Create sheet for this household
            sheet_name = f"Household_{idx + 1}"
            if len(sheet_name) > 31:  # Excel sheet name limit
                sheet_name = sheet_name[:31]
            
            self.worksheet = self.workbook.create_sheet(title=sheet_name)
            
            # Write header block
            self._write_header_block(household, family)
            
            # Write table headers
            self._write_table_headers()
            
            # Write member rows
            self._write_member_rows(family)
        
        # Apply styling to all sheets
        for sheet in self.workbook.worksheets:
            self.worksheet = sheet
            self._apply_styling()
        
        return self.workbook
    
    def _write_header_block(self, household, family):
        """Write the header block with household information."""
        # Header information
        headers = [
            ('Region', ''),
            ('Province', ''),
            ('City/Municipality', ''),
            ('Barangay', household.barangay.name),
            ('Household Address', household.house_number),
            ('No. of Household Members', family.members.count()),
        ]
        
        row_idx = 1
        for label, value in headers:
            self.worksheet.cell(row=row_idx, column=1, value=label)
            self.worksheet.cell(row=row_idx, column=2, value=value)
            row_idx += 1
        
        # Empty row before table
        row_idx += 1
        return row_idx
    
    def _write_table_headers(self):
        """Write the table column headers."""
        headers = [
            'LAST NAME', 'FIRST NAME', 'MIDDLE NAME', 'EXT', 'PLACE OF BIRTH',
            'DATE OF BIRTH', 'AGE', 'SEX', 'CIVIL STATUS', 'CITIZENSHIP',
            'OCCUPATION', 'Indicate if Labor/employed, Unemployed, PWD, OFW, Solo Parent (OSY), Out of School Children (OSC) and/or IP'
        ]
        
        row_idx = 6  # Start after header block
        for col_idx, header in enumerate(headers, start=1):
            self.worksheet.cell(row=row_idx, column=col_idx, value=header)
        
        return row_idx + 1
    
    def _write_member_rows(self, family):
        """Write member rows for the family."""
        members = family.members.all().order_by('first_name', 'last_name')
        row_idx = 7  # Start after table headers
        
        for member in members:
            self.worksheet.cell(row=row_idx, column=1, value=member.last_name or '')
            self.worksheet.cell(row=row_idx, column=2, value=member.first_name or '')
            self.worksheet.cell(row=row_idx, column=3, value=member.middle_name or '')
            self.worksheet.cell(row=row_idx, column=4, value=member.suffix or '')
            self.worksheet.cell(row=row_idx, column=5, value=member.birthplace or '')
            
            # Format birthdate
            if member.birthdate:
                birthdate_cell = self.worksheet.cell(row=row_idx, column=6, value=member.birthdate)
                birthdate_cell.number_format = 'M/D/YYYY'
            else:
                self.worksheet.cell(row=row_idx, column=6, value='')
            
            # Age (calculated)
            self.worksheet.cell(row=row_idx, column=7, value=member.age or '')
            
            # Sex
            sex_display = 'M' if member.sex == 'M' else 'F' if member.sex == 'F' else ''
            self.worksheet.cell(row=row_idx, column=8, value=sex_display)
            
            # Civil status
            civil_status_display = member.get_civil_status_display() if member.civil_status else ''
            self.worksheet.cell(row=row_idx, column=9, value=civil_status_display)
            
            # Citizenship
            self.worksheet.cell(row=row_idx, column=10, value=member.citizenship or '')
            
            # Occupation
            self.worksheet.cell(row=row_idx, column=11, value=member.occupation or '')
            
            # Indicate if column (reconstruct from boolean flags)
            indicate_if_parts = []
            if member.is_pwd:
                indicate_if_parts.append('PWD')
            if member.is_solo_parent:
                indicate_if_parts.append('Solo Parent')
            if member.is_out_of_school_youth:
                indicate_if_parts.append('OSY')
            if member.is_out_of_school_children:
                indicate_if_parts.append('OSC')
            if member.is_indigenous:
                indicate_if_parts.append('IP')
            if member.is_senior_citizen:
                indicate_if_parts.append('Senior Citizen')
            
            indicate_if_text = ', '.join(indicate_if_parts) if indicate_if_parts else ''
            self.worksheet.cell(row=row_idx, column=12, value=indicate_if_text)
            
            row_idx += 1
    
    def _apply_styling(self):
        """Apply styling to make the spreadsheet look like RBI Form A."""
        # Bold font for header labels
        bold_font = Font(bold=True)
        
        # Style header block
        for row_idx in range(1, 6):
            cell = self.worksheet.cell(row=row_idx, column=1)
            cell.font = bold_font
        
        # Style table headers
        header_row = 6
        for col_idx in range(1, 13):
            cell = self.worksheet.cell(row=header_row, column=col_idx)
            cell.font = bold_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        # Set column widths
        column_widths = [20, 20, 20, 10, 25, 15, 8, 8, 15, 15, 25, 50]
        for col_idx, width in enumerate(column_widths, start=1):
            self.worksheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
        
        # Wrap text for the last column
        last_col = openpyxl.utils.get_column_letter(12)
        self.worksheet.column_dimensions[last_col].width = 60
        for row in self.worksheet.iter_rows(min_row=6, max_col=12):
            row[11].alignment = Alignment(wrap_text=True, vertical='top')
