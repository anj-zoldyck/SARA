from urllib import request
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required
from django.forms import ValidationError
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponseForbidden, JsonResponse, HttpResponseRedirect
from django.contrib import messages
from django.utils import timezone
from datetime import date
from django.db.models import Count, Q

from accounts.decorators import session_protected, mswdo_or_staff_required
from accounts.models import User, Barangay
from households.models import Household, Zone, Family, FamilyMember
from programs.models import Program, AidCategory, Assistance
from distribution.models import AidSchedule, AidClaim
from .models import ReportGenerationLog
from .analytics_utils import get_quarterly_report_data, get_beneficiary_list_data
from xhtml2pdf import pisa
from django.template.loader import get_template
from accounts.forms import CreateUserForm
from households.forms import HouseholdForm, FamilyForm, FamilyMemberForm
from programs.forms import ProgramForm, AidCategoryForm, AssistanceForm
# from distribution.forms import AidScheduleForm  # if any

from django.utils.safestring import mark_safe
from distribution.services import get_active_aid_schedule, get_active_schedule
from django.utils.dateparse import parse_datetime
from django_otp.plugins.otp_email.models import EmailDevice
from django.urls import reverse
from django.core.cache import cache
import json

from core.audit_utils import log_action

User = get_user_model()


@login_required
@session_protected
@mswdo_or_staff_required
def aid_reports(request):

    selected_barangay = request.GET.get('barangay')

    schedules = AidSchedule.objects.select_related(
    'assistance', 'assistance__program', 'assistance__aid_category', 'barangay'
        ).all().order_by('-schedule_datetime')

    if selected_barangay:
        schedules = schedules.filter(barangay_id=selected_barangay)

    # Attach claims per schedule
    for sched in schedules:
        sched.claims = AidClaim.objects.filter(
            schedule=sched          #filter by schedule directly
        ).select_related(
            'family',
            'family_member',
            'family__household',
            'family__household__barangay',
            'assistance',
            'assistance__aid_category',
        )

    barangays = Barangay.objects.all()

    return render(request, 'reports/aid_reports.html', {
        'schedules': schedules,
        'barangays': barangays,
        'selected_barangay': selected_barangay,
        'now': timezone.now(),
    })


@login_required
def generate_summary_report(request):
    if request.user.role != 'MSWDO':
        return HttpResponseForbidden("Only MSWDO Admin can generate this report.")
        
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    if not start_date_str or not end_date_str:
        return HttpResponse("Missing date range.", status=400)
        
    from django.utils.dateparse import parse_datetime
    start_date = parse_datetime(start_date_str + "T00:00:00").date()
    end_date = parse_datetime(end_date_str + "T23:59:59").date()
    
    data = get_quarterly_report_data(start_date, end_date)
    
    chart_image = None
    if len(data['months']) > 1:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        import io, base64
        
        fig, ax = plt.subplots(figsize=(8, 4))
        
        months = [m['month_label'] for m in data['months']]
        categories = set()
        for m in data['months']:
            for c in m['categories']:
                categories.add(c['name'])
        categories = list(categories)
        categories.sort()
        
        x = range(len(months))
        width = 0.8 / max(1, len(categories))
        
        for i, cat in enumerate(categories):
            counts = []
            for m in data['months']:
                count = next((c['count'] for c in m['categories'] if c['name'] == cat), 0)
                counts.append(count)
            ax.bar([pos + i*width for pos in x], counts, width, label=cat)
            
        ax.set_xticks([pos + width*(len(categories)-1)/2 for pos in x])
        ax.set_xticklabels(months)
        ax.legend()
        
        plt.tight_layout()
        buf = io.BytesIO()
        plt.savefig(buf, format='png')
        plt.close(fig)
        chart_image = "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode('utf-8')
        
    context = {
        'data': data,
        'generated_at': timezone.now(),
        'chart_image': chart_image
    }
    
    template = get_template('reports/summary_report_pdf.html')
    html = template.render(context)
    
    from django.http import HttpResponse
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Summary_Report.pdf"'
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    if not pisa_status.err:
        report_log = ReportGenerationLog.objects.create(
            report_type='SUMMARY',
            period_label=data['period_label'],
            generated_by=request.user
        )
        log_action(request.user, 'REPORT_GENERATED', target=report_log, description=f"Generated SUMMARY report for {data['period_label']}")
        return response
    return HttpResponse("Error generating PDF", status=500)


@login_required
def generate_beneficiary_list_report(request):
    if request.user.role != 'MSWDO':
        return HttpResponseForbidden("Only MSWDO Admin can generate this report.")
        
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    
    if not start_date_str or not end_date_str:
        return HttpResponse("Missing date range.", status=400)
        
    from django.utils.dateparse import parse_datetime
    start_date = parse_datetime(start_date_str + "T00:00:00").date()
    end_date = parse_datetime(end_date_str + "T23:59:59").date()
    
    data = get_beneficiary_list_data(start_date, end_date)
    
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Beneficiary List"
    
    headers = ["Name", "Barangay", "Zone", "Assistance/Category", "Date Claimed"]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
    column_widths = [30, 25, 20, 35, 20]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
        
    for row_idx, row_data in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=row_data['name'])
        ws.cell(row=row_idx, column=2, value=row_data['barangay'])
        ws.cell(row=row_idx, column=3, value=row_data['zone'])
        ws.cell(row=row_idx, column=4, value=row_data['assistance'])
        claimed_date_str = timezone.localtime(row_data['claimed_at']).strftime('%Y-%m-%d %H:%M')
        ws.cell(row=row_idx, column=5, value=claimed_date_str)
        
    from django.http import HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Beneficiary_List_Report.xlsx"'
    
    wb.save(response)
    wb.close()
    
    from django.utils.dateformat import DateFormat
    period_label = f"{DateFormat(start_date).format('F Y')} to {DateFormat(end_date).format('F Y')}"
    report_log = ReportGenerationLog.objects.create(
        report_type='BENEFICIARY_LIST',
        period_label=period_label,
        generated_by=request.user
    )
    log_action(request.user, 'REPORT_GENERATED', target=report_log, description=f"Generated BENEFICIARY_LIST report for {period_label}")
    
    return response
