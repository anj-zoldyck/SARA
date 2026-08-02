from django.test import TestCase, Client
from django.urls import reverse
from accounts.models import User, Barangay
from households.models import Household, Family, FamilyMember, Zone
from programs.models import Program, AidCategory, Assistance
from distribution.models import AidSchedule, AidClaim
from reports.models import ReportGenerationLog
import datetime
from django.utils import timezone
import openpyxl
from io import BytesIO

class ReportingFeatureTests(TestCase):
    def setUp(self):
        # Create users
        self.mswdo_admin = User.objects.create_user(username='admin', email='admin@test.com', password='password', role='MSWDO')
        self.mswdo_staff = User.objects.create_user(username='staff', email='staff@test.com', password='password', role='MSWDO_STAFF')
        
        self.client = Client()
        
        # Create models
        self.barangay = Barangay.objects.create(name='San Juan')
        self.zone = Zone.objects.create(barangay=self.barangay, name='Zone 1')
        self.household = Household.objects.create(barangay=self.barangay, zone=self.zone, house_number='123 Main St')
        self.family = Family.objects.create(household=self.household, family_name='Dela Cruz')
        self.member = FamilyMember.objects.create(family=self.family, first_name='Juan', last_name='Dela Cruz', sex='M')
        
        self.program = Program.objects.create(name='Assistance Program')
        self.category = AidCategory.objects.create(program=self.program, name='Medical Assistance')
        self.category2 = AidCategory.objects.create(program=self.program, name='Financial Assistance')
        
        self.assistance_family = Assistance.objects.create(program=self.program, aid_category=self.category, beneficiary_type='family')
        self.assistance_individual = Assistance.objects.create(program=self.program, aid_category=self.category2, beneficiary_type='individual')
        
        self.schedule = AidSchedule.objects.create(assistance=self.assistance_family, schedule_datetime=timezone.now(), budget=1000, per_beneficiary_amount=100)
        
        # Add claims
        self.claim1 = AidClaim.objects.create(family=self.family, assistance=self.assistance_family, schedule=self.schedule)
        self.claim1.claimed_at = timezone.now() - datetime.timedelta(days=1)
        self.claim1.save()
        
        self.claim2 = AidClaim.objects.create(family=self.family, assistance=self.assistance_individual, family_member=self.member, schedule=self.schedule)
        self.claim2.claimed_at = self.claim1.claimed_at
        self.claim2.save()
        
        self.start_date = (self.claim1.claimed_at - datetime.timedelta(days=2)).strftime('%Y-%m-%d')
        self.end_date = (self.claim1.claimed_at + datetime.timedelta(days=2)).strftime('%Y-%m-%d')

    def test_permissions(self):
        # Unauthenticated
        response = self.client.get(reverse('generate_summary_report'), {'start_date': self.start_date, 'end_date': self.end_date})
        self.assertEqual(response.status_code, 302) # Redirect to login
        
        # Staff role
        self.client.force_login(self.mswdo_staff)
        response = self.client.get(reverse('generate_summary_report'), {'start_date': self.start_date, 'end_date': self.end_date})
        self.assertEqual(response.status_code, 403)
        
    def test_summary_report_generation(self):
        self.client.force_login(self.mswdo_admin)
        response = self.client.get(reverse('generate_summary_report'), {'start_date': self.start_date, 'end_date': self.end_date})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        
        # Verify generation log
        self.assertTrue(ReportGenerationLog.objects.filter(report_type='SUMMARY', generated_by=self.mswdo_admin).exists())
        
    def test_beneficiary_list_report_generation(self):
        self.client.force_login(self.mswdo_admin)
        response = self.client.get(reverse('generate_beneficiary_list_report'), {'start_date': self.start_date, 'end_date': self.end_date})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # Verify excel content
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        
        # Header row + 2 claims = 3 rows
        self.assertEqual(ws.max_row, 3)
        
        # Read back claim data (Dela Cruz Family)
        row2_name = ws.cell(row=2, column=1).value
        self.assertIn('Dela Cruz', row2_name)
        
        # Verify generation log
        self.assertTrue(ReportGenerationLog.objects.filter(report_type='BENEFICIARY_LIST', generated_by=self.mswdo_admin).exists())

    def test_empty_report_generation(self):
        self.client.force_login(self.mswdo_admin)
        empty_start = (timezone.now() + datetime.timedelta(days=10)).strftime('%Y-%m-%d')
        empty_end = (timezone.now() + datetime.timedelta(days=15)).strftime('%Y-%m-%d')
        
        response = self.client.get(reverse('generate_summary_report'), {'start_date': empty_start, 'end_date': empty_end})
        self.assertEqual(response.status_code, 200) # Should return empty PDF, not error
        
        response = self.client.get(reverse('generate_beneficiary_list_report'), {'start_date': empty_start, 'end_date': empty_end})
        self.assertEqual(response.status_code, 200) # Should return empty Excel
        
        wb = openpyxl.load_workbook(BytesIO(response.content))
        self.assertEqual(wb.active.max_row, 1) # Only header row

    def test_unique_beneficiaries_vs_total_claims(self):
        """
        Specific test for unique beneficiaries vs total claims distinction.
        """
        from reports.analytics_utils import get_quarterly_report_data
        
        start = timezone.now() - datetime.timedelta(days=10)
        end = timezone.now() + datetime.timedelta(days=10)
        
        AidClaim.objects.all().delete()
        
        # Two claims for the exact same family in the same month
        claimA = AidClaim.objects.create(family=self.family, assistance=self.assistance_family, schedule=self.schedule)
        claimA.claimed_at = timezone.now()
        claimA.save()
        
        claimB = AidClaim.objects.create(family=self.family, assistance=self.assistance_family, schedule=self.schedule)
        claimB.claimed_at = timezone.now()
        claimB.save()
        
        data = get_quarterly_report_data(start.date(), end.date())
        month_data = data['months'][0]
        
        self.assertEqual(month_data['total_claims'], 2)
        self.assertEqual(month_data['total_beneficiaries'], 1)
